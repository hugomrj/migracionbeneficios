

import uuid
import openpyxl
import pandas as pd
from flask import current_app as app
from config import db
from app.models import  Beneficio




def migrar_beneficios_excel(file):    
    """Procesa el archivo Excel y migra los datos a la base de datos."""
    
    # Generar un UUID para esta migración
    proceso_id = str(uuid.uuid4().hex)
    
   
    
    # Leer el archivo Excel
    df = pd.read_excel(file, sheet_name=0)  


    """Recorre toda la primera columna y muestra cada valor."""
    if df.empty:
        raise ValueError("El archivo está vacío")


    # Verificar filas ocultas
    if not verificar_filas_columnas_ocultos(file):
        raise ValueError("El archivo contiene filas ocultas.")


    resultado = verificar_anio_y_valor_numerico(file)
    if resultado:        
        anio = resultado
        print(f"El valor de 'AÑO' es: {resultado}")
    else:
        anio = 0        
        raise ValueError("Error en el  registro de AÑO")



    resultado = verificar_mes_y_valor_numerico(file)
    if resultado:        
        mes = resultado
        print(f"El valor de 'MES' es: {resultado}")
    else:
        mes = 0
        raise ValueError("Error en el registro de MES")




    resultado = verificar_objeto_y_valor_numerico(file)
    if resultado:        
        objeto = resultado
        print(f"El valor de 'OBJETO' es: {resultado}")
    else:
        objeto = 0
        raise ValueError("Error en el registro de OBJETO")





    resultado = verificar_concepto_y_valor_numerico(file)
    if resultado:        
        concepto = resultado
        print(f"El valor de 'CONCEPTO' es: {resultado}")
    else:
        concepto = 0
        raise ValueError("Error en el registro de CONCEPTO")


    # Leer el archivo Excel
    df = pd.read_excel(file, sheet_name=0, header=4)
    with app.app_context():
        try:
            for index, row in df.iterrows():

                # Verificar si alguna columna importante tiene NaN y asignar un valor predeterminado si es necesario
                cedula = row.get('CEDULA')
                presupues = row.get('PRESUPUESTADO', 0)
                devengado = row.get('DEVENGADO', 0)
                jubilacion = row.get('JUBILACION', 0)
                liquido = row.get('LIQUIDO', 0)



                # Verificar si el valor de la cédula es NaN
                if pd.isna(cedula):
                    print(f"Fila {index + 1} - Cédula está vacío, omitiendo fila.")
                    continue  # O puedes manejar esto de otra manera si lo prefieres

                # Verificar si el valor de presupues es NaN
                if pd.isna(presupues):
                    presupues = 0
                    print(f"Fila {index + 1} - Presupues está vacío, asignando valor 0.")

                # Verificar si el valor de devengado es NaN
                if pd.isna(devengado):
                    devengado = 0
                    print(f"Fila {index + 1} - Devengado está vacío, asignando valor 0.")

                # Verificar si el valor de jubilacion es NaN
                if pd.isna(jubilacion):
                    jubilacion = 0
                    print(f"Fila {index + 1} - Jubilacion está vacío, asignando valor 0.")

                # Verificar si el valor de liquido es NaN
                if pd.isna(liquido):
                    liquido = 0
                    print(f"Fila {index + 1} - Liquido está vacío, asignando valor 0.")



                # Crear una instancia de Beneficio con los datos del archivo Excel
                beneficio = Beneficio(
                    cedula=cedula,
                    presupues=presupues,
                    devengado=devengado,
                    jubilacion=jubilacion,
                    liquido=liquido,
                    proceso_id=proceso_id,
                    mes=mes,
                    anio=anio,
                    codigo_concepto=concepto,
                    ccargo=0,  # Valor por defecto si no está en el Excel
                    tipo_traba=0  # Valor por defecto si no está en el Excel
                )
                # Agregar la instancia a la sesión
                db.session.add(beneficio)
                
            # Confirmar los cambios
            db.session.commit()
            print(f"Proceso completado con ID: {proceso_id}")

            return proceso_id  

        except Exception as e:
            # Revertir cualquier cambio en caso de error
            db.session.rollback()
            print(f"Error en la migración: {e}")            
            
        
        

def verificar_filas_columnas_ocultos(excel_file):
    """Verifica si hay filas o columnas ocultas en el archivo Excel."""
    try:
        # Cargar el libro de trabajo con openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active  # Obtener la primera hoja activa

        # Verificar filas ocultas
        filas_ocultas = [row.row for row in ws.row_dimensions.values() if row.hidden]
        
        # Verificar columnas ocultas
        columnas_ocultas = [col.column for col in ws.column_dimensions.values() if col.hidden]

        if filas_ocultas or columnas_ocultas:
            return False  # Existen filas o columnas ocultas
        return True  # No hay filas ni columnas ocultas

    except Exception as e:
        print(f"Error al verificar filas o columnas ocultas: {e}")
        return False  # En caso de error, considerar que hay filas o columnas ocultas






def verificar_anio_y_valor_numerico(excel_file):
    """Verifica que el primer registro, primera columna sea 'AÑO' y que el segundo valor sea numérico."""
    try:
        # Cargar el libro de trabajo con openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active  # Obtener la primera hoja activa

        # Verificar que la primera celda contenga "AÑO"
        valor_ano = ws.cell(row=1, column=1).value
        if valor_ano != "AÑO":
            return False  # Si no es "AÑO", retornar False

        # Obtener el valor en la primera fila, segunda columna
        valor_columna_2 = ws.cell(row=1, column=2).value
        if not isinstance(valor_columna_2, (int, float)):
            return False  # Si no es un valor numérico, retornar False

        # Retornar el valor numérico si todo es correcto
        return valor_columna_2

    except Exception as e:
        print(f"Error al verificar el archivo: {e}")
        return False







def verificar_mes_y_valor_numerico(excel_file):
    """Verifica que el segundo registro, primera columna sea 'MES' y que el valor en la segunda columna sea numérico."""
    try:
        # Cargar el libro de trabajo con openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active  # Obtener la primera hoja activa

        # Verificar que la celda (2,1) contenga "MES"
        valor_mes = ws.cell(row=2, column=1).value
        if valor_mes != "MES":
            return False  # Si no es "MES", retornar False

        # Obtener el valor en la segunda fila, segunda columna
        valor_columna_2 = ws.cell(row=2, column=2).value
        if not isinstance(valor_columna_2, (int, float)):
            return False  # Si no es un valor numérico, retornar False

        # Retornar el valor numérico si todo es correcto
        return valor_columna_2

    except Exception as e:
        print(f"Error al verificar el archivo: {e}")
        return False








def verificar_objeto_y_valor_numerico(excel_file):
    """Verifica que el tercer registro, primera columna sea 'OBJETO' y que el valor en la segunda columna sea numérico."""
    try:
        # Cargar el libro de trabajo con openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active  # Obtener la primera hoja activa

        # Verificar que la celda (3,1) contenga "OBJETO"
        valor_objeto = ws.cell(row=3, column=1).value
        if valor_objeto != "OBJETO":
            return False  # Si no es "OBJETO", retornar False

        # Obtener el valor en la tercera fila, segunda columna
        valor_columna_2 = ws.cell(row=3, column=2).value
        if not isinstance(valor_columna_2, (int, float)):
            return False  # Si no es un valor numérico, retornar False

        # Retornar el valor numérico si todo es correcto
        return valor_columna_2

    except Exception as e:
        print(f"Error al verificar el archivo: {e}")
        return False






def verificar_concepto_y_valor_numerico(excel_file):
    """Verifica que el cuarto registro, primera columna sea 'CONCEPTO' y que el valor en la segunda columna sea numérico."""
    try:
        # Cargar el libro de trabajo con openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active  # Obtener la primera hoja activa

        # Verificar que la celda (4,1) contenga "CONCEPTO"
        valor_concepto = ws.cell(row=4, column=1).value
        if valor_concepto != "CONCEPTO":
            return False  # Si no es "CONCEPTO", retornar False

        # Obtener el valor en la cuarta fila, segunda columna
        valor_columna_2 = ws.cell(row=4, column=2).value
        if not isinstance(valor_columna_2, (int, float)):
            return False  # Si no es un valor numérico, retornar False

        # Retornar el valor numérico si todo es correcto
        return valor_columna_2

    except Exception as e:
        print(f"Error al verificar el archivo: {e}")
        return False







            
            
            
            
            
            
            
            